import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\utkuesin.kurucu\Downloads\15_DC15_2025_09_13.xlsx')

for ws in wb.worksheets:
    sheet_name = ws.title
    print(f"\n{'='*60}")
    print(f"SAYFA: {sheet_name}")
    print(f"{'='*60}")
    
    # Tarife sütunlarını bul
    tarife_cols = []
    for i in range(4, 30):
        cell = ws.cell(5, i)
        if cell.value and str(cell.value).startswith('T'):
            tarife_cols.append((i, cell.value))
    
    print(f"Tarife columns: {len(tarife_cols)}")
    
    # Hareket satırlarını bul
    hareket_rows = []
    for i in range(8, ws.max_row + 1):
        cell = ws.cell(i, 2)
        if cell.value in ['Kalkış', 'Dönüş']:
            hareket_rows.append((i, cell.value))
    
    print(f"Hareket rows: {len(hareket_rows)}")
    print(f"Rows: {[f'{h}({r})' for r, h in hareket_rows]}")
    
    # Son Dönüş satırını bul
    last_donus_row = [r for r, h in hareket_rows if h == 'Dönüş']
    last_donus = last_donus_row[-1] if last_donus_row else None
    
    # Son Dönüş'ün altındaki satırda veri var mı kontrol et
    skip_last_donus = False
    if last_donus:
        has_any_data = False
        for col, tarife in tarife_cols:
            cell_below = ws.cell(last_donus + 1, col)
            # Sadece değer kontrolü yap (formül de olabilir)
            if cell_below.value:
                has_any_data = True
                break
        
        skip_last_donus = not has_any_data
        print(f"Son Dönüş (Row {last_donus}): Altında veri {'VAR - İŞLENECEK' if has_any_data else 'YOK - ATLANACAK'}")
    
    # Merged cell ranges
    merged_ranges = list(ws.merged_cells.ranges)
    
    kalkis_count = 0
    donus_count = 0
    
    for row_num, hareket in hareket_rows:
        # SON DÖNÜŞ'Ü ATLA (sadece altında veri yoksa)
        if skip_last_donus and hareket == 'Dönüş' and row_num == last_donus:
            continue
        
        row_added = 0
        
        for col, tarife in tarife_cols:
            cell = ws.cell(row_num, col)
            
            if not cell.value:
                continue
            
            # Beyaz font
            if cell.font and cell.font.color and hasattr(cell.font.color, 'theme'):
                if cell.font.color.theme == 0 and (cell.font.color.tint or 0) >= 0:
                    continue
            
            # Dönüş için altını kontrol et
            if hareket == 'Dönüş':
                cell_below = ws.cell(row_num + 1, col)
                
                if not cell_below.value:
                    continue
                
                if cell_below.font and cell_below.font.color and hasattr(cell_below.font.color, 'theme'):
                    if cell_below.font.color.theme == 0 and (cell_below.font.color.tint or 0) >= 0:
                        continue
                
                # Merged cell kontrolü
                cell_coord = cell_below.coordinate
                if any(cell_coord in merged_range for merged_range in merged_ranges):
                    continue
            
            row_added += 1
        
        if hareket == 'Kalkış':
            kalkis_count += row_added
        else:
            donus_count += row_added
    
    print(f"\nKalkış: {kalkis_count}")
    print(f"Dönüş: {donus_count}")
    print(f"TOPLAM: {kalkis_count + donus_count}")
