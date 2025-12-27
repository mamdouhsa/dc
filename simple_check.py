#!/usr/bin/env python3
import openpyxl
import sys

file_path = sys.argv[1]
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb.active

print(f"\n{'='*70}")
print(f"AC05 - DÖNÜŞ SATIRLARI ANALİZİ")
print(f"{'='*70}\n")

# Header bul
header_row = None
for r in range(1, 21):
    for c in range(4, 31):
        val = ws.cell(r, c).value
        if val and str(val).startswith('T') and len(str(val)) > 1:
            header_row = r
            break
    if header_row:
        break

print(f"Header row: {header_row}\n")

# Dönüş satırları bul
donus_rows = []
start_row = header_row + 2 if header_row else 7
for r in range(start_row, 50):
    cell = ws.cell(r, 2)
    if cell.value and str(cell.value).strip() == 'Dönüş':
        donus_rows.append(r)
        
        # Merged cell kontrolü
        if cell.coordinate in ws.merged_cells:
            print(f"⚠️  Satır {r}: B sütunu merged cell içinde!")

print(f"\n{'='*70}")
print(f"DÖNÜŞ SATIRLARINDAKİ TARİFE HÜCRELERİ")
print(f"{'='*70}\n")

for donus_row in donus_rows:
    print(f"\n📍 SATIR {donus_row} - DÖNÜŞ\n")
    
    for col in range(4, 20):  # D-S sütunları (T01-T16 civarı)
        cell = ws.cell(donus_row, col)
        
        if not cell.value:
            continue
        
        col_letter = openpyxl.utils.get_column_letter(col)
        
        # Font color
        font_info = "No font"
        if cell.font and cell.font.color:
            if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                rgb = cell.font.color.rgb
                font_info = f"RGB: {rgb}"
            elif hasattr(cell.font.color, 'theme') and cell.font.color.theme is not None:
                font_info = f"THEME: {cell.font.color.theme}"
            elif hasattr(cell.font.color, 'indexed') and cell.font.color.indexed is not None:
                font_info = f"INDEXED: {cell.font.color.indexed}"
            else:
                font_info = f"Color object: {cell.font.color}"
        
        # Value
        val = str(cell.value)[:30]
        
        print(f"  {col_letter}{donus_row}: {val}")
        print(f"      Font: {font_info}")
        
        # Formül mü?
        if str(cell.value).startswith('='):
            print(f"      ⚙️  FORMÜL HÜCRESİ")
        
        print()

print(f"\n{'='*70}")
print(f"ÖNEMLİ: Hangi hücreler BEYAZ yazılı?")
print(f"Excel'de bu satırlara bak ve bana söyle:")
for dr in donus_rows:
    print(f"  - Satır {dr} (Dönüş)")
print(f"{'='*70}\n")
