import openpyxl
import sys

# Excel dosyasını aç
file_path = r'C:\Users\utkuesin.kurucu\Downloads\22_ML22_2025_10_12.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"📊 Dosya: {file_path}")
    print(f"📋 Toplam sayfa sayısı: {len(wb.sheetnames)}")
    print(f"📄 Sayfalar: {', '.join(wb.sheetnames)}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"📄 SAYFA: {sheet_name}")
        print(f"{'='*60}")
        
        # İlk 25 satırı kontrol et
        print("\n🔍 İlk 25 satırın başlıkları:")
        for row_num in range(1, min(26, ws.max_row + 1)):
            row_values = []
            for col_num in range(1, min(31, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    row_values.append(f"[{openpyxl.utils.get_column_letter(col_num)}] {str(cell.value)[:30]}")
            
            if row_values:
                print(f"  Satır {row_num}: {' | '.join(row_values[:10])}")
        
        # İlk satırdaki tüm değerleri göster
        print(f"\n📌 İlk satır (Satır 1) - TÜM SÜTUNLAR:")
        first_row = ws[1]
        for idx, cell in enumerate(first_row, 1):
            if cell.value:
                print(f"  {openpyxl.utils.get_column_letter(idx)}{1}: '{cell.value}'")
        
        # İkinci satırdaki tüm değerleri göster
        print(f"\n📌 İkinci satır (Satır 2) - TÜM SÜTUNLAR:")
        if ws.max_row >= 2:
            second_row = ws[2]
            for idx, cell in enumerate(second_row, 1):
                if cell.value:
                    print(f"  {openpyxl.utils.get_column_letter(idx)}{2}: '{cell.value}'")

except Exception as e:
    print(f"❌ HATA: {e}")
    import traceback
    traceback.print_exc()
