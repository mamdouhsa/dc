import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\utkuesin.kurucu\Downloads\15_DC15_2025_09_13.xlsx')
ws = wb.active

print("=== Searching for T01, T02... headers ===")
for row_num in range(1, 21):
    row_values = []
    for col in range(4, 25):
        cell = ws.cell(row_num, col)
        if cell.value and str(cell.value).strip().startswith('T'):
            row_values.append(f"{cell.column_letter}{row_num}:{cell.value}")
    if row_values:
        print(f"Row {row_num}: {', '.join(row_values)}")

print("\n=== Row 17 details (Kalkış) ===")
print(f"B17: {ws.cell(17, 2).value}")
for col in range(4, 19):
    cell = ws.cell(17, col)
    col_letter = cell.column_letter
    font_theme = cell.font.color.theme if cell.font and cell.font.color and hasattr(cell.font.color, 'theme') else None
    font_rgb = cell.font.color.rgb if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None
    print(f"{col_letter}17: {cell.value} | theme={font_theme} | rgb={font_rgb}")
