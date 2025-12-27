import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\utkuesin.kurucu\Downloads\15_DC15_2025_09_13.xlsx')
ws = wb.active

# Merged cell range'lerini al
merged_ranges = list(ws.merged_cells.ranges)
print(f"Merged cells: {merged_ranges}\n")

hareket_rows = [(8,'Kalkış'),(12,'Dönüş'),(17,'Kalkış'),(21,'Dönüş'),(26,'Kalkış'),(30,'Dönüş'),(35,'Kalkış'),(39,'Dönüş'),(44,'Kalkış'),(48,'Dönüş')]

# Son Dönüş satırını bul
last_donus_row = [r for r, h in hareket_rows if h == 'Dönüş'][-1]

kalkis_count = 0
donus_count = 0

for row_num, hareket in hareket_rows:
    # SON DÖNÜŞ'Ü TAMAMEN ATLA
    if hareket == 'Dönüş' and row_num == last_donus_row:
        print(f"{hareket} Row {row_num}: ATLANDI (son Dönüş)")
        continue
    
    row_added = 0
    row_skipped = 0
    
    for col in range(4, 19):  # D to R (T01 to T15)
        cell = ws.cell(row_num, col)
        
        # Boş hücre
        if not cell.value:
            row_skipped += 1
            continue
        
        # Beyaz font kontrolü
        if cell.font and cell.font.color and hasattr(cell.font.color, 'theme'):
            if cell.font.color.theme == 0 and (cell.font.color.tint or 0) >= 0:
                row_skipped += 1
                continue
        
        # Dönüş için altındaki hücreyi kontrol et (SON DÖNÜŞ HARİÇ)
        if hareket == 'Dönüş' and row_num != last_donus_row:
            cell_below = ws.cell(row_num + 1, col)
            
            # Altı boş
            if not cell_below.value:
                row_skipped += 1
                continue
            
            # Altı beyaz
            if cell_below.font and cell_below.font.color and hasattr(cell_below.font.color, 'theme'):
                if cell_below.font.color.theme == 0 and (cell_below.font.color.tint or 0) >= 0:
                    row_skipped += 1
                    continue
            
            # Altı merged cell mi?
            cell_coord = cell_below.coordinate
            is_merged = any(cell_coord in merged_range for merged_range in merged_ranges)
            if is_merged:
                row_skipped += 1
                continue
        
        row_added += 1
    
    print(f"{hareket} Row {row_num}: {row_added} added, {row_skipped} skipped")
    
    if hareket == 'Kalkış':
        kalkis_count += row_added
    else:
        donus_count += row_added

print(f"\n=== TOTAL ===")
print(f"Kalkış: {kalkis_count}")
print(f"Dönüş: {donus_count}")
print(f"GRAND TOTAL: {kalkis_count + donus_count}")
