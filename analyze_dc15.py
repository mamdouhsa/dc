import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\utkuesin.kurucu\Downloads\15_DC15_2025_09_13.xlsx')
ws = wb.active

# Tarife sütunlarını bul
tarife_cols = []
for i in range(4, 30):
    cell = ws.cell(5, i)
    if cell.value and str(cell.value).startswith('T'):
        tarife_cols.append((i, cell.value))

print(f"Found {len(tarife_cols)} tarife columns: {[t[1] for t in tarife_cols]}\n")

# Hareket satırlarını bul
hareket_rows = []
for i in range(8, ws.max_row + 1):
    cell = ws.cell(i, 2)  # B sütunu
    if cell.value in ['Kalkış', 'Dönüş']:
        hareket_rows.append((i, cell.value))

print(f"Found {len(hareket_rows)} hareket rows: {hareket_rows}\n")

# Her hareket satırı için kontrol
for row_num, hareket in hareket_rows:
    print(f"\n=== {hareket} (Row {row_num}) ===")
    
    for col_num, tarife_name in tarife_cols:
        cell = ws.cell(row_num, col_num)
        col_letter = openpyxl.utils.get_column_letter(col_num)
        
        # Değer kontrolü
        if not cell.value:
            print(f"  SKIP (empty): {col_letter}{row_num} ({tarife_name})")
            continue
        
        # Font rengi kontrolü
        font_color = None
        if cell.font and cell.font.color:
            if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                font_color = f"rgb={cell.font.color.rgb}"
            elif hasattr(cell.font.color, 'theme'):
                theme = cell.font.color.theme
                tint = cell.font.color.tint if hasattr(cell.font.color, 'tint') else 0
                font_color = f"theme={theme}, tint={tint:.2f}"
                
                # Beyaz kontrolü
                if theme == 0 and tint >= 0:
                    print(f"  SKIP (white font): {col_letter}{row_num} ({tarife_name}) = {cell.value} | {font_color}")
                    continue
        
        # Dönüş için altındaki hücreyi kontrol et
        if hareket == 'Dönüş':
            cell_below = ws.cell(row_num + 1, col_num)
            if not cell_below.value:
                print(f"  SKIP (below empty): {col_letter}{row_num} ({tarife_name}) = {cell.value}")
                continue
            
            # Altındaki hücrenin beyaz olup olmadığını kontrol et
            if cell_below.font and cell_below.font.color and hasattr(cell_below.font.color, 'theme'):
                below_theme = cell_below.font.color.theme
                below_tint = cell_below.font.color.tint if hasattr(cell_below.font.color, 'tint') else 0
                if below_theme == 0 and below_tint >= 0:
                    print(f"  SKIP (below white): {col_letter}{row_num} ({tarife_name}) = {cell.value}")
                    continue
        
        # Bu noktaya geldiyse, hücre eklenebilir
        print(f"  OK: {col_letter}{row_num} ({tarife_name}) = {cell.value}")

