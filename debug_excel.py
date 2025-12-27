#!/usr/bin/env python3
"""
Excel Hata Ayıklama Script
Dosya yapısı sorunlarını tespit etmek için
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os


def debug_excel(file_path):
    """Excel dosyasını detaylı şekilde debug et"""
    
    if not os.path.exists(file_path):
        print(f"❌ Dosya bulunamadı: {file_path}")
        return
    
    print(f"\n{'='*70}")
    print(f"🔧 EXCEL DOSYA DEBUG")
    print(f"{'='*70}\n")
    print(f"Dosya: {file_path}\n")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"✅ Dosya yüklendi")
        print(f"📋 Sheet: {ws.title}")
        print(f"📊 Boyut: {ws.max_row} satır x {ws.max_column} sütun\n")
        
        # ========== STEP 1: B SÜTUNUNU KONTROL ET ==========
        print(f"\n{'='*70}")
        print(f"STEP 1: B SÜTUNUNU KONTROL ET")
        print(f"{'='*70}\n")
        
        b_column_values = []
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            cell = ws.cell(row_idx, 2)
            if cell.value:
                b_column_values.append((row_idx, cell.value))
        
        if b_column_values:
            print(f"✅ B sütununda {len(b_column_values)} dolu hücre bulundu:\n")
            for row_idx, value in b_column_values:
                is_hareket = value in ['Kalkış', 'Dönüş']
                mark = "✅" if is_hareket else "⚠️"
                print(f"{mark} Satır {row_idx:3d}: '{value}'")
            
            hareket_count = len([v for r, v in b_column_values if v in ['Kalkış', 'Dönüş']])
            if hareket_count == 0:
                print(f"\n❌ PROBLEM: 'Kalkış' veya 'Dönüş' satırı bulunamadı!")
                print(f"   Lütfen B sütunundaki değerleri kontrol edin.")
                print(f"   Yazım: Tam olarak 'Kalkış' veya 'Dönüş' olmalı")
        else:
            print(f"❌ B sütununda hiç veri yok!")
        
        # ========== STEP 2: D SÜTUNÜNDEN BAŞLAYAN BAŞLIKLARI KONTROL ET ==========
        print(f"\n{'='*70}")
        print(f"STEP 2: D SÜTÜNÜNDEN BAŞLAYAN BAŞLIKLARI KONTROL ET")
        print(f"{'='*70}\n")
        
        print("İlk 20 satırda T başlayan başlıkları arıyor...\n")
        
        tarife_headers = {}
        for row_idx in range(1, min(21, ws.max_row + 1)):
            for col_idx in range(4, ws.max_column + 1):  # D = 4
                cell = ws.cell(row_idx, col_idx)
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip().upper()
                    if value.startswith('T') and len(value) >= 2:
                        try:
                            int(value[1:])
                            if row_idx not in tarife_headers:
                                tarife_headers[row_idx] = []
                            tarife_headers[row_idx].append({
                                'name': value,
                                'col': col_idx,
                                'letter': get_column_letter(col_idx)
                            })
                        except ValueError:
                            pass
        
        if tarife_headers:
            print(f"✅ {sum(len(v) for v in tarife_headers.values())} Tarife başlığı bulundu:\n")
            for row_idx in sorted(tarife_headers.keys()):
                headers = tarife_headers[row_idx]
                print(f"  Satır {row_idx}: {len(headers)} başlık")
                for h in headers[:10]:
                    print(f"    - {h['name']} ({h['letter']} sütunu)")
                if len(headers) > 10:
                    print(f"    ... ve {len(headers) - 10} tane daha")
        else:
            print(f"❌ PROBLEM: T01, T02, T03... başlıkları bulunamadı!")
            print(f"   Lütfen Excel dosyanızda D sütününden başlayan T başlıkları olup olmadığını kontrol edin.")
            
            # İlk satırları göster
            print(f"\n   İlk satırın D-Z sütunları:")
            for col_idx in range(4, min(27, ws.max_column + 1)):
                cell = ws.cell(1, col_idx)
                col_letter = get_column_letter(col_idx)
                value = cell.value if cell.value else "(boş)"
                print(f"     {col_letter}: {value}")
        
        # ========== STEP 3: MERGED HÜCRELER ==========
        print(f"\n{'='*70}")
        print(f"STEP 3: MERGED HÜCRELER (Birleştirilmiş Hücreler)")
        print(f"{'='*70}\n")
        
        merged_ranges = list(ws.merged_cells.ranges)
        if merged_ranges:
            print(f"⚠️  {len(merged_ranges)} merged hücre bulundu:\n")
            for i, merged_range in enumerate(merged_ranges[:10], 1):
                print(f"  {i}. {merged_range}")
            if len(merged_ranges) > 10:
                print(f"\n  ... ve {len(merged_ranges) - 10} tane daha")
            
            print(f"\n⚠️  NOT: Merged hücreler parsing sırasında sorun yaratabilir.")
            print(f"   Bu normaldir ve kod bu durumu ele almalıdır.")
        else:
            print(f"✅ Merged hücre yok - İyi haber!\n")
        
        # ========== STEP 4: ÖRNEK VERİ ==========
        print(f"\n{'='*70}")
        print(f"STEP 4: ÖRNEK VERİ (İlk 10 satır, D-M sütunları)")
        print(f"{'='*70}\n")
        
        print("    ", end="")
        for col_idx in range(4, min(14, ws.max_column + 1)):
            col_letter = get_column_letter(col_idx)
            print(f"{col_letter:^8}", end="")
        print()
        
        for row_idx in range(1, min(11, ws.max_row + 1)):
            print(f"R{row_idx:2d}:", end="")
            for col_idx in range(4, min(14, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value)[:6] if cell.value else ""
                print(f"{value:^8}", end="")
            print()
        
        print()
        
        # ========== ÖZET ==========
        print(f"\n{'='*70}")
        print(f"ÖZET")
        print(f"{'='*70}\n")
        
        issues = []
        
        hareket_count = len([v for r, v in b_column_values if v in ['Kalkış', 'Dönüş']])
        if hareket_count == 0:
            issues.append("❌ B sütununda 'Kalkış' veya 'Dönüş' satırı yok")
        else:
            print(f"✅ {hareket_count} Hareket satırı bulundu")
        
        tarife_count = sum(len(v) for v in tarife_headers.values())
        if tarife_count == 0:
            issues.append("❌ T01, T02... tarife başlıkları yok")
        else:
            print(f"✅ {tarife_count} Tarife başlığı bulundu")
        
        if issues:
            print("\n🚨 SORUNLAR:\n")
            for issue in issues:
                print(f"  {issue}")
            print("\n📖 ÇÖZÜM:")
            print("  1. Excel dosyasını açın")
            print("  2. B sütunundaki satırlara 'Kalkış' veya 'Dönüş' yazın")
            print("  3. D sütününden başlayarak T01, T02, T03... başlıkları ekleyin")
            print("  4. Test scriptini tekrar çalıştırın")
        else:
            print("\n✅ Dosya yapısı sorunsuz görünüyor!")
        
        print(f"\n{'='*70}\n")
        
    except Exception as e:
        print(f"❌ Hata: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("❌ Hata: Excel dosya yolu gerekli")
        print("\nKullanım:")
        print('  python debug_excel.py "dosya.xlsx"')
        sys.exit(1)
    
    file_path = sys.argv[1]
    debug_excel(file_path)
