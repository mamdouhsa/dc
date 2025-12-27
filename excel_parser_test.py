#!/usr/bin/env python3
"""
Excel Parser Test Script
Test amaçlı: Excel dosyalarını parse edip B sütununda "Kalkış"/"Dönüş" 
ve D+ sütunlarında T01, T02 başlıklarını ve saat verilerini bulmak
"""

import openpyxl
from openpyxl.utils import get_column_letter
import sys
import os
from pathlib import Path


class ExcelParserTester:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.data = {
            'tarife_columns': [],
            'hareket_rows': [],
            'parsed_data': []
        }
        
    def load_workbook(self):
        """Excel dosyasını yükle"""
        if not os.path.exists(self.file_path):
            print(f"❌ Hata: Dosya bulunamadı: {self.file_path}")
            return False
            
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            print(f"✅ Dosya yüklendi: {self.file_path}")
            print(f"📋 Sheet adı: {self.worksheet.title}")
            return True
        except Exception as e:
            print(f"❌ Dosya yükleme hatası: {e}")
            return False
    
    def print_sheet_info(self):
        """Sheet bilgilerini yazdır"""
        if not self.worksheet:
            return
            
        print(f"\n${'='*60}")
        print(f"📊 SHEET BİLGİSİ")
        print(f"${'='*60}")
        print(f"Toplam satır: {self.worksheet.max_row}")
        print(f"Toplam sütun: {self.worksheet.max_column}")
        
    def print_first_rows(self, num_rows=15):
        """İlk satırları yazdır"""
        if not self.worksheet:
            return
            
        print(f"\n${'='*60}")
        print(f"📋 İLK {num_rows} SATIR (A-P sütunları)")
        print(f"${'='*60}\n")
        
        for row_idx in range(1, min(num_rows + 1, self.worksheet.max_row + 1)):
            print(f"Satır {row_idx:2d}:", end="")
            for col_idx in range(1, min(17, self.worksheet.max_column + 1)):
                cell = self.worksheet.cell(row_idx, col_idx)
                col_letter = get_column_letter(col_idx)
                value = cell.value
                
                if value is not None:
                    value_str = str(value)[:15]
                    print(f" | {col_letter}: {value_str}", end="")
            print()
        print()
    
    def find_tarife_columns(self):
        """T01, T02 gibi tarife sütunlarını bul"""
        print(f"\n${'='*60}")
        print(f"🔍 TARIFE SÜTUNLARINI ARIYORUM")
        print(f"${'='*60}\n")
        
        # İlk 20 satırda ara
        for row_idx in range(1, min(21, self.worksheet.max_row + 1)):
            for col_idx in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row_idx, col_idx)
                
                if cell.value and isinstance(cell.value, str):
                    value = cell.value.strip().upper()
                    
                    # T01, T02, T03 gibi başlıkları bul
                    if value.startswith('T') and len(value) >= 2:
                        try:
                            num = int(value[1:])
                            col_letter = get_column_letter(col_idx)
                            
                            # Zaten ekli mi?
                            if not any(t['name'] == value for t in self.data['tarife_columns']):
                                self.data['tarife_columns'].append({
                                    'name': value,
                                    'col_index': col_idx,
                                    'col_letter': col_letter,
                                    'header_row': row_idx
                                })
                                print(f"✅ BULUNDU: {col_letter} (sütun {col_idx}) = {value} (satır {row_idx})")
                        except ValueError:
                            pass
        
        if self.data['tarife_columns']:
            print(f"\n✅ {len(self.data['tarife_columns'])} Tarife sütunu bulundu")
            # Sıralandığını kontrol et
            self.data['tarife_columns'].sort(key=lambda x: x['col_index'])
        else:
            print(f"\n❌ Tarife sütunu bulunamadı!")
        
        return len(self.data['tarife_columns']) > 0
    
    def find_hareket_rows(self):
        """B sütununda "Kalkış"/"Dönüş" satırlarını bul"""
        print(f"\n${'='*60}")
        print(f"🔍 B SÜTUNUNDA 'KALKIŞ'/'DÖNÜŞ' ARIYORUM")
        print(f"${'='*60}\n")
        
        for row_idx in range(1, self.worksheet.max_row + 1):
            cell = self.worksheet.cell(row_idx, 2)  # B sütunu = 2
            
            if cell.value:
                value = cell.value.strip()
                
                if value in ['Kalkış', 'Dönüş']:
                    self.data['hareket_rows'].append({
                        'row': row_idx,
                        'hareket': value
                    })
                    print(f"✅ BULUNDU: Satır {row_idx} = '{value}'")
        
        if self.data['hareket_rows']:
            print(f"\n✅ {len(self.data['hareket_rows'])} Hareket satırı bulundu")
        else:
            print(f"\n❌ Hareket satırı bulunamadı!")
        
        return len(self.data['hareket_rows']) > 0
    
    def parse_data(self):
        """Verileri parse et"""
        print(f"\n${'='*60}")
        print(f"📊 VERİLER PARSE EDİLİYOR")
        print(f"${'='*60}\n")
        
        if not self.data['tarife_columns'] or not self.data['hareket_rows']:
            print("❌ Tarife sütunları veya Hareket satırları bulunamadı!")
            return False
        
        for hareket_info in self.data['hareket_rows']:
            row_idx = hareket_info['row']
            hareket = hareket_info['hareket']
            
            print(f"\n📍 Satır {row_idx}: {hareket}")
            
            for tarife_info in self.data['tarife_columns']:
                col_idx = tarife_info['col_index']
                tarife_name = tarife_info['name']
                col_letter = tarife_info['col_letter']
                
                cell = self.worksheet.cell(row_idx, col_idx)
                value = cell.value
                
                if value is not None:
                    # Zaman formatını kontrol et
                    cell_ref = f"{col_letter}{row_idx}"
                    time_value = str(value).strip()
                    
                    # Excel time formatı (decimal)
                    if isinstance(value, float) and 0 < value < 1:
                        total_seconds = round(value * 86400)
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        time_value = f"{hours:02d}:{minutes:02d}:00"
                        print(f"  ✅ {cell_ref} ({tarife_name}): {value} -> {time_value} (decimal format)")
                    # HH:MM format
                    elif ':' in time_value and time_value.count(':') == 1:
                        time_value = f"{time_value}:00"
                        print(f"  ✅ {cell_ref} ({tarife_name}): {time_value}")
                    # HH:MM:SS format
                    elif ':' in time_value and time_value.count(':') == 2:
                        print(f"  ✅ {cell_ref} ({tarife_name}): {time_value}")
                    else:
                        print(f"  ⚠️ {cell_ref} ({tarife_name}): Tanımlanamayan format: '{time_value}'")
                        continue
                    
                    self.data['parsed_data'].append({
                        'tarife': tarife_name,
                        'hareket': hareket,
                        'tarife_saati': time_value,
                        'onaylanan': None,
                        'durum': None,
                        'plaka': None
                    })
        
        print(f"\n✅ {len(self.data['parsed_data'])} kayıt parse edildi")
        return len(self.data['parsed_data']) > 0
    
    def print_merged_cells(self):
        """Merged hücreleri yazdır"""
        if not self.worksheet:
            return
        
        merged_ranges = self.worksheet.merged_cells.ranges
        
        if merged_ranges:
            print(f"\n${'='*60}")
            print(f"🔗 MERGED HÜCRELER ({len(merged_ranges)} adet)")
            print(f"${'='*60}\n")
            
            for i, merged_range in enumerate(list(merged_ranges)[:15]):
                print(f"  {merged_range}")
            
            if len(merged_ranges) > 15:
                print(f"  ... ve {len(merged_ranges) - 15} tane daha")
        else:
            print(f"\n✅ Merged hücre yok")
    
    def print_summary(self):
        """Özet yazdır"""
        print(f"\n${'='*60}")
        print(f"📊 ÖZET")
        print(f"${'='*60}\n")
        
        print(f"✅ Tarife sütunları: {len(self.data['tarife_columns'])} adet")
        if self.data['tarife_columns']:
            for t in self.data['tarife_columns'][:5]:
                print(f"   - {t['name']} ({t['col_letter']} sütunu, satır {t['header_row']})")
            if len(self.data['tarife_columns']) > 5:
                print(f"   ... ve {len(self.data['tarife_columns']) - 5} tane daha")
        
        print(f"\n✅ Hareket satırları: {len(self.data['hareket_rows'])} adet")
        if self.data['hareket_rows']:
            for h in self.data['hareket_rows']:
                print(f"   - Satır {h['row']}: {h['hareket']}")
        
        print(f"\n✅ Parse edilen kayıtlar: {len(self.data['parsed_data'])} adet")
        if self.data['parsed_data']:
            print(f"\n   İlk 5 kayıt:")
            for i, record in enumerate(self.data['parsed_data'][:5], 1):
                print(f"   {i}. {record['tarife']} | {record['hareket']} | {record['tarife_saati']}")
            if len(self.data['parsed_data']) > 5:
                print(f"   ... ve {len(self.data['parsed_data']) - 5} kayıt daha")
        
        print(f"\n{'='*60}\n")
        
        # Başarı durumu
        if len(self.data['parsed_data']) > 0:
            print("✅ ✅ ✅ TEST BAŞARILI - Tüm veriler parse edildi!")
            return True
        else:
            print("❌ TEST BAŞARISIZ - Veri parse edilemedi!")
            return False
    
    def run_test(self):
        """Tüm testleri çalıştır"""
        print(f"\n{'='*60}")
        print(f"🚀 EXCEL PARSER TEST")
        print(f"{'='*60}\n")
        
        if not self.load_workbook():
            return False
        
        self.print_sheet_info()
        self.print_first_rows()
        self.print_merged_cells()
        
        if not self.find_tarife_columns():
            return False
        
        if not self.find_hareket_rows():
            return False
        
        if not self.parse_data():
            return False
        
        return self.print_summary()


def main():
    if len(sys.argv) < 2:
        print("❌ Hata: Excel dosya yolu gerekli")
        print("\nKullanım:")
        print("  python excel_parser_test.py <dosya_yolu>")
        print("\nÖrnek:")
        print('  python excel_parser_test.py "49_TCD49A_2025_10_14.xlsx"')
        print('  python excel_parser_test.py "/path/to/dosya.xlsx"')
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    tester = ExcelParserTester(file_path)
    success = tester.run_test()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
