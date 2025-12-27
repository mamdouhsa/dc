#!/usr/bin/env python3
"""
Beyaz Font Kontrolü
Excel dosyasındaki beyaz yazılı hücreleri tespit eder
"""

import openpyxl
from openpyxl.styles import Color
import sys

if len(sys.argv) < 2:
    print("❌ Kullanım: python check_white_font.py file.xlsx")
    sys.exit(1)

file_path = sys.argv[1]

wb = openpyxl.load_workbook(file_path)
ws = wb.active

print(f"\n{'='*70}")
print(f"BEYAZ FONT KONTROLÜ")
print(f"{'='*70}\n")
print(f"Dosya: {file_path}")
print(f"Sheet: {ws.title}\n")

# Önce header row bul (T01, T02...)
header_row = None
for row_idx in range(1, 21):
    for col_idx in range(4, 31):
        cell = ws.cell(row_idx, col_idx)
        if cell.value and str(cell.value).strip().startswith('T') and str(cell.value).strip()[1:3].isdigit():
            header_row = row_idx
            break
    if header_row:
        break

print(f"Header row: {header_row}\n")

# Dönüş satırlarını bul
donus_rows = []
for row_idx in range(header_row + 2 if header_row else 7, 50):
    cell = ws.cell(row_idx, 2)  # B sütunu
    if cell.value and str(cell.value).strip() == 'Dönüş':
        donus_rows.append(row_idx)

print(f"{'='*70}")
print(f"DÖNÜŞ SATIRLARI: {len(donus_rows)} adet")
print(f"{'='*70}\n")

for donus_row in donus_rows:
    print(f"\n📍 SATIR {donus_row} - DÖNÜŞ")
    print(f"{'-'*70}\n")
    
    # D sütunundan itibaren (tarife sütunları)
    for col_idx in range(4, 31):
        cell = ws.cell(donus_row, col_idx)
        
        if not cell.value:
            continue
        
        # Font bilgisi
        font = cell.font
        fill = cell.fill
        
        # Değer
        value = str(cell.value)[:20]
        
        # Font rengi
        font_color = None
        font_color_type = None
        
        if font and font.color:
            if font.color.rgb:
                font_color = str(font.color.rgb)
                font_color_type = "RGB"
            elif font.color.theme is not None:
                font_color = f"theme={font.color.theme}"
                font_color_type = "THEME"
            elif font.color.indexed is not None:
                font_color = f"indexed={font.color.indexed}"
                font_color_type = "INDEXED"
        
        # Fill rengi
        fill_color = None
        if fill and hasattr(fill, 'fgColor') and fill.fgColor:
            if fill.fgColor.rgb:
                fill_color = str(fill.fgColor.rgb)
        
        # Beyaz kontrolü
        is_white = False
        white_reason = ""
        
        if font_color_type == "RGB":
            # RGB beyaz: FFFFFF veya FFFFFFFF (son 6 karakter)
            if font_color and font_color[-6:].upper() == 'FFFFFF':
                is_white = True
                white_reason = f"RGB={font_color}"
        elif font_color_type == "THEME":
            # Theme 1 genelde beyaz
            if "theme=1" in font_color or "theme=0" in font_color:
                is_white = True
                white_reason = font_color
        
        mark = "⚪ BEYAZ" if is_white else "✅ NORMAL"
        
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"{mark} {col_letter}{donus_row}: '{value}'")
        if font_color:
            print(f"     Font: {font_color_type} - {font_color}")
        if fill_color:
            print(f"     Fill: {fill_color}")
        if is_white:
            print(f"     ⚠️  ATLANMALI: {white_reason}")
        print()

print(f"\n{'='*70}")
print(f"ÖZET")
print(f"{'='*70}\n")
print(f"✅ Toplam {len(donus_rows)} Dönüş satırı bulundu")
print(f"📋 Yukarıdaki listede ⚪ BEYAZ işaretli hücreler ATLANmalı\n")
